import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import openpyxl
import io

# ── 頁面設定 ──
st.set_page_config(
    page_title="富強醫材 設備維修 KPI 儀表板",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── 自訂樣式 ──
st.markdown("""
<style>
    .main { padding: 1rem 2rem; }
    .block-container { padding-top: 1rem; }
    .stMetric { background: #f5f4f0; border-radius: 8px; padding: 0.75rem 1rem; }
    .stMetric label { font-size: 12px; color: #888780; }
    .stMetric [data-testid="metric-container"] { background: #f5f4f0; border-radius: 8px; }
    div[data-testid="column"] { gap: 0.5rem; }
    .series-header { font-size: 14px; font-weight: 600; color: #1a1a18; margin: 1rem 0 0.5rem; }
    .badge {
        display: inline-block; padding: 2px 10px; border-radius: 4px;
        font-size: 12px; font-weight: 600;
    }
    .badge-red    { background: #fcebeb; color: #a32d2d; }
    .badge-green  { background: #e2efda; color: #1f7a4d; }
    .badge-teal   { background: #e1f5ee; color: #0f6e56; }
    .badge-amber  { background: #fff2cc; color: #7b4f00; }
    .badge-blue   { background: #e6f1fb; color: #185fa5; }
    h1 { font-size: 22px !important; color: #1A3A5C !important; }
    h2 { font-size: 16px !important; color: #2E6DA4 !important; }
    h3 { font-size: 14px !important; }
    .upload-area { background: #fff; border: 2px dashed #b4b2a9; border-radius: 12px;
                   padding: 2rem; text-align: center; margin: 2rem auto; max-width: 600px; }
</style>
""", unsafe_allow_html=True)

# ── 常數 ──
SRS_NAMES = ['押出機','單機型壓鑄成型機','自動射出成型機','液態矽膠射出成型機','無廢料射出成型機','後射式矽膠射出成型機']
SRS_COLORS = ['#378ADD','#1D9E75','#BA7517','#D85A30','#7F77DD','#D4537E']
COLOR_MAP = dict(zip(SRS_NAMES, SRS_COLORS))
SRS_SHORT = {
    '押出機': '押出機',
    '單機型壓鑄成型機': '壓鑄',
    '自動射出成型機': '自動射出',
    '液態矽膠射出成型機': '液態矽膠',
    '無廢料射出成型機': '無廢料',
    '後射式矽膠射出成型機': '後射式矽膠',
}
MONTH_OFFSETS = [2,7,12,17,22,27,32,37,42,47,52,57]

def get_series(name):
    if not name: return None
    n = str(name)
    for s in SRS_NAMES:
        if s[:3] in n: return s
    return None

def status_badge(label):
    cls = {'需重點改善':'badge-red','可靠度高':'badge-green','表現卓越':'badge-teal',
           '待觀察':'badge-amber','維修高效':'badge-blue','持續改善中':'badge-amber','趨於穩定':'badge-blue'}
    c = cls.get(label,'badge-blue')
    return f'<span class="badge {c}">{label}</span>'

def compute_status(mtbf_vals, mttr_vals):
    valid_b = [v for v in mtbf_vals if v is not None]
    valid_t = [v for v in mttr_vals if v > 0]
    if not valid_b: return '表現卓越'
    avg_b = sum(valid_b)/len(valid_b)
    avg_t = sum(valid_t)/len(valid_t) if valid_t else 0
    improving = len(valid_b)>=2 and valid_b[-1]>valid_b[0]
    if avg_b >= 400: return '可靠度高'
    if avg_b >= 200: return '趨於穩定' if improving else '待觀察'
    if avg_t <= 10 and avg_t > 0: return '維修高效'
    if improving: return '持續改善中'
    return '需重點改善' if avg_b < 50 else '待觀察'

# ── Excel 解析 ──
@st.cache_data
def parse_excel(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 找 KPI 工作表
    kpi_sheet = next((s for s in wb.sheetnames if '關鍵設備績效指標' in s), None)
    if not kpi_sheet:
        raise ValueError("找不到「關鍵設備績效指標」工作表")
    ws = wb[kpi_sheet]
    kpi_rows = list(ws.iter_rows(values_only=True))

    # 動態偵測月份列
    month_row_idx, month_cols = -1, []
    for ri, row in enumerate(kpi_rows[:20]):
        found = [(ci, str(v).strip()) for ci, v in enumerate(row or []) if v and str(v).strip().endswith('月') and str(v).strip()[:-1].isdigit()]
        if found:
            month_row_idx = ri
            month_cols = found
            break
    if month_row_idx == -1:
        raise ValueError("無法偵測月份列")

    # 年份與標準時數
    month_row = kpi_rows[month_row_idx]
    year_label = next((str(v).strip() for v in (month_row or []) if v and '年' in str(v)), '2026年')
    std_row = kpi_rows[month_row_idx + 1] or []

    months = []
    for ci, label in month_cols:
        std_h = 0
        try: std_h = float(std_row[ci]) if std_row[ci] else 0
        except: pass
        if std_h > 0:  # 只保留有標準時數的月份（有實際資料）
            months.append({'label': year_label+label, 'short': label, 'offset': ci, 'stdHrs': std_h})

    # 機台資料起始列
    machine_start = month_row_idx + 3
    for ri in range(month_row_idx+1, min(len(kpi_rows), month_row_idx+10)):
        row = kpi_rows[ri] or []
        if any(str(c).startswith('M-') for c in row if c):
            machine_start = ri; break

    def pn(v):
        if v is None or str(v).strip() == '無故障': return 0
        try: return float(str(v).replace(',',''))
        except: return 0

    machines = []
    for row in kpi_rows[machine_start:]:
        if not row or not row[0] or not row[1]: continue
        name = str(row[0]).strip()
        code = str(row[1]).strip()
        if not code.startswith('M-'): continue
        mdata = []
        for m in months:
            mc = m['offset']
            f = int(pn(row[mc])) if len(row) > mc else 0
            r = round(pn(row[mc+1]),2) if len(row) > mc+1 else 0
            d = round(pn(row[mc+2]),2) if len(row) > mc+2 else 0
            braw = row[mc+3] if len(row) > mc+3 else None
            mtbf = None if (braw is None or str(braw).strip()=='無故障' or f==0) else round(float(braw),2)
            mttr = round(pn(row[mc+4]),2) if (len(row) > mc+4 and f>0) else 0
            mdata.append({'f':f,'r':r,'d':d,'mtbf':mtbf,'mttr':mttr})
        machines.append({'name':name,'code':code,'months':mdata})

    # 月報統計
    monthly_stats = []
    for m in months:
        sname = next((s for s in wb.sheetnames if m['short'] in s), None)
        total, comp = 0, 0
        if sname:
            ws2 = wb[sname]
            for ri2, row2 in enumerate(ws2.iter_rows(min_row=1, max_row=6, values_only=True)):
                row2 = list(row2) if row2 else []
                # 第3列（index2）= 報修件數，col 2
                if ri2 == 2 and len(row2) > 2:
                    v = pn(row2[2])
                    if 30 <= v <= 500: total = int(v)
                # 第4列（index3）= 完成率，col 5
                if ri2 == 3 and len(row2) > 5:
                    v = pn(row2[5])
                    if 0.5 < v < 1.0: comp = v
                # 備援：掃描找完成率
                if comp == 0 or total == 0:
                    for cell in (row2 or []):
                        v = pn(cell)
                        if 0.5 < v < 1.0 and comp == 0: comp = v
                        if 30 <= v <= 500 and v == int(v) and total == 0: total = int(v)
        monthly_stats.append({'totalRepairs': total, 'completionRate': comp})

    return months, machines, monthly_stats

# ── 系列彙總 ──
def aggregate_series(machines, months):
    result = {}
    for name in SRS_NAMES:
        ms = [m for m in machines if m['name'] == name]
        by_month = []
        for mi, month in enumerate(months):
            f = sum(m['months'][mi]['f'] for m in ms)
            r = round(sum(m['months'][mi]['r'] for m in ms), 2)
            d = round(sum(m['months'][mi]['d'] for m in ms), 2)
            std = month['stdHrs']
            mtbf = round((std-d)/f, 2) if f > 0 else None
            mttr = round(r/f, 2) if f > 0 else 0
            by_month.append({'f':f,'r':r,'d':d,'mtbf':mtbf,'mttr':mttr})
        result[name] = by_month
    return result

# ══════════════════════════════
#  메인 앱
# ══════════════════════════════
def main():
    # ── 標題 ──
    st.markdown("## 🏭 富強醫材股份有限公司 — 設備維修 KPI 儀表板")

    # ── 上傳 ──
    if 'data' not in st.session_state:
        st.markdown("""
        <div class="upload-area">
            <div style="font-size:48px;margin-bottom:12px">📊</div>
            <div style="font-size:18px;font-weight:600;color:#1a1a18;margin-bottom:8px">請上傳設備維修統計 Excel 檔案</div>
            <div style="font-size:13px;color:#888780">支援格式：.xlsx　需包含「關鍵設備績效指標」工作表</div>
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader("", type=['xlsx'], label_visibility='collapsed')
        if uploaded:
            with st.spinner("解析資料中..."):
                try:
                    months, machines, monthly_stats = parse_excel(uploaded.read())
                    st.session_state['data'] = (months, machines, monthly_stats)
                    st.session_state['filename'] = uploaded.name
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ 解析失敗：{e}")
        return

    months, machines, monthly_stats = st.session_state['data']
    series_agg = aggregate_series(machines, months)
    month_labels = [m['label'] for m in months]
    month_shorts = [m['short'] for m in months]

    # ── 頂部資訊列 ──
    col_info, col_btn = st.columns([4,1])
    with col_info:
        st.caption(f"📁 {st.session_state.get('filename','')}　　📅 {month_labels[0]} – {month_labels[-1]}　　富強醫材股份有限公司")
    with col_btn:
        if st.button("↩ 重新上傳"):
            del st.session_state['data']
            st.rerun()

    st.divider()

    # ── 分頁 ──
    tab1, tab2 = st.tabs(["📊 機台系列", "🔍 個別機台"])

    # ══ 機台系列 ══
    with tab1:
        c1, c2 = st.columns(2)
        with c1:
            sel_month_s = st.selectbox("月份", ["全部月份"] + month_labels, key='s_month')
        with c2:
            sel_series_s = st.selectbox("機台系列", ["全部機台系列"] + SRS_NAMES, key='s_series')

        # 篩選月份索引
        if sel_month_s == "全部月份":
            midxs = list(range(len(months)))
        else:
            midxs = [month_labels.index(sel_month_s)]

        # 篩選系列
        filtered_series = [sel_series_s] if sel_series_s != "全部機台系列" else SRS_NAMES

        # KPI 計算
        total_f = sum(series_agg[s][mi]['f'] for s in filtered_series for mi in midxs)
        total_r = sum(series_agg[s][mi]['r'] for s in filtered_series for mi in midxs)
        total_d = sum(series_agg[s][mi]['d'] for s in filtered_series for mi in midxs)
        total_rep = sum(monthly_stats[mi]['totalRepairs'] for mi in midxs) if sel_series_s == "全部機台系列" else total_f
        avg_comp = sum(monthly_stats[mi]['completionRate'] for mi in midxs) / len(midxs) if midxs else 0

        # KPI Cards
        k1,k2,k3,k4,k5 = st.columns(5)
        k1.metric("設備報修總數", f"{total_rep} 件", f"{len(midxs)} 個月累計")
        k2.metric("故障次數", f"{total_f} 次", f"{len(filtered_series)} 個系列")
        k3.metric("維修工時", f"{total_r:.1f} h", "含等待+作業時間")
        delta_d = f"停機 {total_d:.1f}h"
        k4.metric("停工工時", f"{total_d:.1f} h", delta_d, delta_color="inverse")
        comp_delta = "✓ 達標" if avg_comp >= 0.95 else "△ 需改善"
        k5.metric("平均完成率", f"{avg_comp*100:.1f}%", comp_delta)

        # ── Row 1: MTBF + MTTR ──
        xlabels = [month_shorts[i] for i in midxs]
        c1, c2 = st.columns(2)

        with c1:
            fig = go.Figure()
            for name in filtered_series:
                color = COLOR_MAP[name]
                y = [series_agg[name][i]['mtbf'] for i in midxs]
                fig.add_trace(go.Scatter(
                    x=xlabels, y=y, name=SRS_SHORT.get(name, name),
                    mode='lines+markers', line=dict(color=color, width=2),
                    marker=dict(size=7), connectgaps=True
                ))
            fig.update_layout(
                title=dict(text="MTBF 趨勢（hr / 次）", font_size=14, x=0),
                height=320, margin=dict(t=50,b=40,l=50,r=120),
                legend=dict(orientation='v', x=1.01, y=1, font_size=10,
                            bgcolor='rgba(255,255,255,0.8)', bordercolor='#eae8e1', borderwidth=1),
                yaxis_title="MTBF (h)", plot_bgcolor='white', paper_bgcolor='white',
                yaxis=dict(gridcolor='#eae8e1'), xaxis=dict(gridcolor='#eae8e1')
            )
            st.plotly_chart(fig, use_container_width=True)
            st.caption("＊ MTBF（平均無故障間隔時間）— 數值越高代表機台越可靠")

        with c2:
            fig2 = go.Figure()
            for name in filtered_series:
                color = COLOR_MAP[name]
                y = [series_agg[name][i]['mttr'] if series_agg[name][i]['f']>0 else None for i in midxs]
                fig2.add_trace(go.Scatter(
                    x=xlabels, y=y, name=SRS_SHORT.get(name, name),
                    mode='lines+markers', line=dict(color=color, width=2),
                    marker=dict(size=7), connectgaps=True
                ))
            fig2.update_layout(
                title=dict(text="MTTR 趨勢（hr / 次）", font_size=14, x=0),
                height=320, margin=dict(t=50,b=40,l=50,r=120),
                legend=dict(orientation='v', x=1.01, y=1, font_size=10,
                            bgcolor='rgba(255,255,255,0.8)', bordercolor='#eae8e1', borderwidth=1),
                yaxis_title="MTTR (h)", plot_bgcolor='white', paper_bgcolor='white',
                yaxis=dict(gridcolor='#eae8e1'), xaxis=dict(gridcolor='#eae8e1')
            )
            st.plotly_chart(fig2, use_container_width=True)
            st.caption("＊ MTTR（平均修復時間）— 數值越低代表維修效率越高")

        # ── Row 2: 故障 + 維修 + 停工 ──
        c1, c2, c3 = st.columns(3)

        with c1:
            fig3 = go.Figure()
            for name in filtered_series:
                color = COLOR_MAP[name]
                y = [series_agg[name][i]['f'] for i in midxs]
                fig3.add_trace(go.Bar(x=xlabels, y=y, name=SRS_SHORT.get(name,name), marker_color=color, opacity=0.85))
            fig3.update_layout(
                title="故障次數（月別）", height=300, margin=dict(t=40,b=30,l=40,r=10),
                barmode='group',
                legend=dict(orientation='h', y=-0.25, font_size=10, xanchor='center', x=0.5),
                plot_bgcolor='white',
                xaxis=dict(gridcolor='#eae8e1'), yaxis=dict(gridcolor='#eae8e1')
            )
            st.plotly_chart(fig3, use_container_width=True)

        with c2:
            fig4 = go.Figure()
            for name in filtered_series:
                color = COLOR_MAP[name]
                y = [series_agg[name][i]['r'] for i in midxs]
                fig4.add_trace(go.Scatter(x=xlabels, y=y, name=SRS_SHORT.get(name,name), mode='lines+markers',
                    line=dict(color=color, width=2), marker=dict(size=5), connectgaps=True))
            fig4.update_layout(
                title="維修工時（hr，月別）", height=300, margin=dict(t=40,b=30,l=40,r=10),
                legend=dict(orientation='h', y=-0.25, font_size=10, xanchor='center', x=0.5),
                yaxis_title="hr", plot_bgcolor='white',
                xaxis=dict(gridcolor='#eae8e1'), yaxis=dict(gridcolor='#eae8e1')
            )
            st.plotly_chart(fig4, use_container_width=True)

        with c3:
            fig5 = go.Figure()
            for name in filtered_series:
                color = COLOR_MAP[name]
                y = [series_agg[name][i]['d'] for i in midxs]
                fig5.add_trace(go.Bar(x=xlabels, y=y, name=SRS_SHORT.get(name,name), marker_color=color, opacity=0.85))
            fig5.update_layout(
                title="停工工時（hr，月別）", height=300, margin=dict(t=40,b=30,l=40,r=10),
                barmode='group',
                legend=dict(orientation='h', y=-0.25, font_size=10, xanchor='center', x=0.5),
                plot_bgcolor='white',
                xaxis=dict(gridcolor='#eae8e1'), yaxis=dict(gridcolor='#eae8e1')
            )
            st.plotly_chart(fig5, use_container_width=True)

        # ── 系列績效表 ──
        st.markdown("#### 系列績效總覽")
        rows = []
        for name in filtered_series:
            data = [series_agg[name][i] for i in midxs]
            f_tot = sum(d['f'] for d in data)
            r_tot = round(sum(d['r'] for d in data), 1)
            d_tot = round(sum(d['d'] for d in data), 1)
            vb = [d['mtbf'] for d in data if d['mtbf'] is not None]
            vt = [d['mttr'] for d in data if d['f'] > 0]
            avg_b = f"{sum(vb)/len(vb):.1f} h" if vb else "無故障"
            avg_t = f"{sum(vt)/len(vt):.1f} h" if vt else "—"
            st_label = compute_status([d['mtbf'] for d in data], [d['mttr'] for d in data])
            rows.append({'機台系列': name, '故障次數': f_tot,
                         '維修工時(h)': r_tot, '停工工時(h)': d_tot,
                         '平均 MTBF': avg_b, '平均 MTTR': avg_t, '狀態': st_label})
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True,
            column_config={"狀態": st.column_config.TextColumn(width="medium")})

    # ══ 個別機台 ══
    with tab2:
        st_sub = st.radio("檢視角度", ["工時 & 故障", "MTBF & MTTR"], horizontal=True, key='m_sub')
        c1, c2, c3 = st.columns(3)
        with c1:
            sel_month_m = st.selectbox("月份", ["全部月份"]+month_labels, key='m_month')
        with c2:
            sel_series_m = st.selectbox("機台系列", ["全部機台系列"]+SRS_NAMES, key='m_series')
        with c3:
            hide_zero = st.checkbox("隱藏零故障機台")

        midxs_m = list(range(len(months))) if sel_month_m=="全部月份" else [month_labels.index(sel_month_m)]
        mlist = [m for m in machines if sel_series_m=="全部機台系列" or m['name']==sel_series_m]
        if hide_zero:
            mlist = [m for m in mlist if any(m['months'][i]['f']>0 for i in midxs_m)]

        total_f = sum(m['months'][i]['f'] for m in mlist for i in midxs_m)
        total_r = round(sum(m['months'][i]['r'] for m in mlist for i in midxs_m), 1)
        total_d = round(sum(m['months'][i]['d'] for m in mlist for i in midxs_m), 1)
        avg_comp_m = sum(monthly_stats[i]['completionRate'] for i in midxs_m)/len(midxs_m) if midxs_m else 0

        k1,k2,k3,k4,k5 = st.columns(5)
        k1.metric("顯示機台數", f"{len(mlist)} 台", f"共 {len([m for m in machines if sel_series_m=='全部機台系列' or m['name']==sel_series_m])} 台")
        k2.metric("故障次數合計", f"{total_f} 次", f"{len(midxs_m)} 個月")
        k3.metric("維修工時合計", f"{total_r} h", "含等待+作業時間")
        k4.metric("停工工時合計", f"{total_d} h", f"停機 {total_d}h", delta_color="inverse")
        k5.metric("平均完成率", f"{avg_comp_m*100:.1f}%", "✓ 達標" if avg_comp_m>=0.95 else "△ 需改善")

        # ── 工時 & 故障 子頁 ──
        if st_sub == "工時 & 故障":
            ranked = sorted(
                [{'code':m['code'],'name':m['name'],
                  'totalF':sum(m['months'][i]['f'] for i in midxs_m),
                  'totalR':round(sum(m['months'][i]['r'] for i in midxs_m),1),
                  'totalD':round(sum(m['months'][i]['d'] for i in midxs_m),1)}
                 for m in mlist],
                key=lambda x: -x['totalF']
            )[:15]
            ranked = [r for r in ranked if r['totalF']>0 or r['totalR']>0]

            if ranked:
                c1, c2 = st.columns(2)
                with c1:
                    fig6 = go.Figure(go.Bar(
                        x=[r['code'] for r in ranked], y=[r['totalF'] for r in ranked],
                        marker_color=[COLOR_MAP.get(r['name'],'#888') for r in ranked],
                        text=[r['totalF'] for r in ranked], textposition='outside'
                    ))
                    fig6.update_layout(title="累計故障次數 TOP 15", height=320,
                        margin=dict(t=40,b=80,l=40,r=10), plot_bgcolor='white',
                        xaxis=dict(tickangle=45,gridcolor='#eae8e1'), yaxis=dict(gridcolor='#eae8e1'))
                    st.plotly_chart(fig6, use_container_width=True)

                with c2:
                    fig7 = go.Figure()
                    fig7.add_trace(go.Bar(
                        x=[r['code'] for r in ranked], y=[r['totalR'] for r in ranked],
                        name='維修工時', marker_color='#378ADD', opacity=0.8
                    ))
                    fig7.add_trace(go.Bar(
                        x=[r['code'] for r in ranked], y=[r['totalD'] for r in ranked],
                        name='停工工時', marker_color='#E24B4A', opacity=0.8
                    ))
                    fig7.update_layout(title="維修工時 vs 停工工時 TOP 15", height=320,
                        barmode='stack', margin=dict(t=40,b=80,l=40,r=10), plot_bgcolor='white',
                        legend=dict(orientation='h', yanchor='bottom', y=1.02),
                        xaxis=dict(tickangle=45,gridcolor='#eae8e1'), yaxis=dict(gridcolor='#eae8e1',title='h'))
                    st.plotly_chart(fig7, use_container_width=True)

            # 明細表
            st.markdown("#### 個別機台明細")
            xlabels_m = [month_shorts[i] for i in midxs_m]
            tbl_rows = []
            for m in mlist:
                row = {'機台名稱': m['name'], '機台編號': m['code']}
                cumF, cumR, cumD = 0, 0, 0
                for mi in midxs_m:
                    d = m['months'][mi]
                    lbl = month_shorts[mi]
                    row[f'{lbl} 故障'] = d['f'] if d['f']>0 else 0
                    row[f'{lbl} 維修(h)'] = d['r'] if d['r']>0 else 0
                    row[f'{lbl} 停工(h)'] = d['d'] if d['d']>0 else 0
                    cumF+=d['f']; cumR+=d['r']; cumD+=d['d']
                if len(midxs_m)>1:
                    row['累計故障'] = cumF
                    row['維修工時(h)'] = round(cumR,1)
                    row['停工工時(h)'] = round(cumD,1)
                tbl_rows.append(row)
            df_m = pd.DataFrame(tbl_rows)
            st.dataframe(df_m, use_container_width=True, hide_index=True)

        # ── MTBF & MTTR 子頁 ──
        else:
            has_fault = [m for m in mlist if any(m['months'][i]['f']>0 for i in midxs_m)]

            if has_fault:
                def avg_val(m, field):
                    if field == 'mtbf':
                        v = [m['months'][i]['mtbf'] for i in midxs_m if m['months'][i]['mtbf'] is not None]
                    else:
                        v = [m['months'][i]['mttr'] for i in midxs_m if m['months'][i]['f']>0]
                    return round(sum(v)/len(v),1) if v else 0

                ranked_kpi = sorted(
                    [{'code':m['code'],'name':m['name'],
                      'avgMTBF':avg_val(m,'mtbf'),'avgMTTR':avg_val(m,'mttr')}
                     for m in has_fault], key=lambda x: -x['avgMTBF']
                )[:15]

                c1, c2, c3 = st.columns(3)
                with c1:
                    fig8 = go.Figure(go.Bar(
                        x=[r['code'] for r in ranked_kpi], y=[r['avgMTBF'] for r in ranked_kpi],
                        marker_color=[COLOR_MAP.get(r['name'],'#888') for r in ranked_kpi], opacity=0.85
                    ))
                    fig8.update_layout(title="MTBF 分佈（有故障機台）", height=300,
                        margin=dict(t=40,b=80,l=40,r=10), plot_bgcolor='white',
                        xaxis=dict(tickangle=45,gridcolor='#eae8e1'), yaxis=dict(gridcolor='#eae8e1',title='h'))
                    st.plotly_chart(fig8, use_container_width=True)

                with c2:
                    sorted_mttr = sorted(ranked_kpi, key=lambda x: x['avgMTTR'])
                    colors_mttr = ['#1f7a4d' if r['avgMTTR']<=5 else '#7b4f00' if r['avgMTTR']<=15 else '#a32d2d' for r in sorted_mttr]
                    fig9 = go.Figure(go.Bar(
                        x=[r['code'] for r in sorted_mttr], y=[r['avgMTTR'] for r in sorted_mttr],
                        marker_color=colors_mttr, opacity=0.85
                    ))
                    fig9.update_layout(title="MTTR 比較（↑ 需改善）", height=300,
                        margin=dict(t=40,b=80,l=40,r=10), plot_bgcolor='white',
                        xaxis=dict(tickangle=45,gridcolor='#eae8e1'), yaxis=dict(gridcolor='#eae8e1',title='h'))
                    st.plotly_chart(fig9, use_container_width=True)

                with c3:
                    fig10 = go.Figure()
                    for r in ranked_kpi:
                        fig10.add_trace(go.Scatter(
                            x=[r['avgMTBF']], y=[r['avgMTTR']],
                            mode='markers+text', name=r['code'],
                            text=[r['code']], textposition='top center',
                            marker=dict(size=10, color=COLOR_MAP.get(r['name'],'#888')),
                            showlegend=False
                        ))
                    fig10.update_layout(
                        title="MTBF vs MTTR 散佈圖", height=300,
                        margin=dict(t=40,b=30,l=40,r=10), plot_bgcolor='white',
                        xaxis=dict(title="MTBF (h) → 越高越可靠", gridcolor='#eae8e1'),
                        yaxis=dict(title="MTTR (h) → 越低越好", gridcolor='#eae8e1')
                    )
                    st.plotly_chart(fig10, use_container_width=True)

            # MTBF/MTTR 明細表
            st.markdown("#### 個別機台 MTBF / MTTR 月別明細")
            tbl_rows2 = []
            for m in mlist:
                row = {'機台名稱': m['name'], '機台編號': m['code']}
                for mi in midxs_m:
                    d = m['months'][mi]
                    lbl = month_shorts[mi]
                    row[f'{lbl} 故障'] = d['f']
                    row[f'{lbl} MTBF(h)'] = d['mtbf'] if d['mtbf'] else '—'
                    row[f'{lbl} MTTR(h)'] = d['mttr'] if d['f']>0 else '—'
                if len(midxs_m)>1:
                    vb = [m['months'][i]['mtbf'] for i in midxs_m if m['months'][i]['mtbf'] is not None]
                    vt = [m['months'][i]['mttr'] for i in midxs_m if m['months'][i]['f']>0]
                    row['平均 MTBF(h)'] = round(sum(vb)/len(vb),1) if vb else '—'
                    row['平均 MTTR(h)'] = round(sum(vt)/len(vt),1) if vt else '—'
                tbl_rows2.append(row)
            df_kpi = pd.DataFrame(tbl_rows2)
            st.dataframe(df_kpi, use_container_width=True, hide_index=True)

    st.caption("富強醫材股份有限公司　設備課　© 2026")

main()
